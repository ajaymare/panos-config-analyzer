"""Generate Excel report for SD-WAN Advisor results."""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from report.styles import (
    BLUE, WHITE, DARK_TEXT, BORDER_COLOR, GRAY,
    header_font, header_fill, header_align,
    data_font, data_align, thin_border, alt_fill,
    title_font, subtitle_font, kpi_label_font,
    kpi_value_font, style_kpi_cell, auto_width,
)

ORANGE = 'FA582D'
LIGHT_ORANGE = 'FDE8E0'
TEAL = '00C0A3'
LIGHT_TEAL = 'E6F9F5'
GREEN = '28A745'
LIGHT_GREEN = 'D4EDDA'
PURPLE = '7D3C98'

panos_fill = PatternFill(start_color=LIGHT_ORANGE, end_color=LIGHT_ORANGE, fill_type='solid')
prisma_fill = PatternFill(start_color=LIGHT_TEAL, end_color=LIGHT_TEAL, fill_type='solid')
section_font = Font(name='Calibri', size=12, bold=True, color=BLUE)


def _write_section_header(ws, row, text, col_count):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = section_font
    cell.alignment = Alignment(vertical='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    return row + 1


def _write_table(ws, start_row, headers, rows):
    """Write a table with headers and data rows. Returns next row."""
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for ri, row_data in enumerate(rows):
        r = start_row + 1 + ri
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if ri % 2 == 1:
                cell.fill = alt_fill

    return start_row + 1 + len(rows) + 1


def generate_advisor_report(result: dict, output_dir: str) -> str:
    """Generate the SD-WAN Advisor Excel report.

    Returns path to the generated .xlsx file.
    """
    wb = Workbook()

    _build_executive_summary(wb, result)
    _build_scoring_sheet(wb, result)
    _build_comparison_sheet(wb, result)
    if result.get('competitive_displacement'):
        _build_competitive_sheet(wb, result)
    _build_next_steps_sheet(wb, result)

    filename = f'SDWAN_Advisor_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_executive_summary(wb, result):
    ws = wb.active
    ws.title = 'Executive Summary'

    inputs = result['inputs']
    rec_label = result['rec_label']
    confidence = result['confidence']
    panos_score = result['panos_score']
    prisma_score = result['prisma_score']

    # Title
    row = 1
    cell = ws.cell(row=row, column=1, value='SD-WAN Advisor — Executive Summary')
    cell.font = title_font
    cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1

    cell = ws.cell(row=row, column=1, value=f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}')
    cell.font = Font(name='Calibri', size=10, italic=True, color=GRAY)
    row += 2

    # KPI row
    kpi_data = [
        ('Recommendation', rec_label),
        ('Confidence', f'{int(confidence * 100)}%'),
        ('PAN-OS Score', f'{panos_score:.0f}/100'),
        ('Prisma Score', f'{prisma_score:.0f}/100'),
    ]
    for ci, (label, value) in enumerate(kpi_data):
        label_cell = ws.cell(row=row, column=ci + 1, value=label)
        label_cell.font = kpi_label_font
        label_cell.alignment = Alignment(horizontal='center')

        value_cell = ws.cell(row=row + 1, column=ci + 1, value=value)
        value_cell.font = kpi_value_font
        value_cell.alignment = Alignment(horizontal='center')

        # Highlight recommendation
        if ci == 0:
            fill_color = ORANGE if result['recommendation'] == 'panos' else TEAL
            value_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            value_cell.font = Font(name='Calibri', size=14, bold=True, color=WHITE)

    row += 3

    # Summary text
    cell = ws.cell(row=row, column=1, value=result['rec_summary'])
    cell.font = Font(name='Calibri', size=11)
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 50
    row += 2

    # Customer profile
    from .engine import COMPETITOR_LABELS
    existing_labels = {'yes_panorama': 'Yes — Panorama', 'yes_scm': 'Yes — SCM', 'no': 'No'}
    security_labels = {'full_ngfw': 'Full NGFW', 'cloud_delivered': 'Cloud Security', 'basic': 'Basic Firewall'}
    mgmt_labels = {'on_prem': 'On-Premises', 'cloud': 'Cloud-Managed', 'no_preference': 'No Preference'}
    priority_labels = {
        'cost': 'Cost Optimization', 'sase': 'SASE / Zero Trust',
        'rapid_deploy': 'Rapid Deployment', 'leverage_investment': 'Leverage Investment',
        'advanced_threat': 'Advanced Threat Prevention', 'iot_security': 'IoT Security',
        'multi_cloud': 'Multi-Cloud',
    }

    row = _write_section_header(ws, row, 'Customer Profile', 4)
    profile_rows = [
        ('Existing PA Firewalls', existing_labels.get(inputs.get('existing_pa', ''), 'N/A')),
        ('Competitor', COMPETITOR_LABELS.get(inputs.get('competitor', ''), 'N/A')),
        ('Branch Sites', inputs.get('branch_count', 'N/A')),
        ('Hub Sites', inputs.get('hub_count', 'N/A')),
        ('Security Requirement', security_labels.get(inputs.get('security', ''), 'N/A')),
        ('Management Preference', mgmt_labels.get(inputs.get('management', ''), 'N/A')),
        ('Priorities', ', '.join(priority_labels.get(p, p) for p in inputs.get('priorities', [])) or 'None'),
    ]
    row = _write_table(ws, row, ['Parameter', 'Value'], profile_rows)

    auto_width(ws)
    ws.column_dimensions['A'].width = max(ws.column_dimensions['A'].width, 25)


def _build_scoring_sheet(wb, result):
    ws = wb.create_sheet('Scoring Breakdown')

    row = 1
    cell = ws.cell(row=row, column=1, value='Scoring Breakdown by Category')
    cell.font = title_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 2

    headers = ['Category', 'Weight', 'PAN-OS Score', 'Prisma Score', 'Rationale']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    row += 1

    from .engine import CATEGORIES
    for key, label, weight in CATEGORIES:
        cs = result['category_scores'][key]
        for ci, val in enumerate([label, f'x{weight}', cs['panos'], cs['prisma'], cs.get('rationale', '')], 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left' if ci in (1, 5) else 'center',
                                       vertical='center', wrap_text=(ci == 5))
            cell.border = thin_border
            # Highlight winning score
            if ci == 3 and cs['panos'] > cs['prisma']:
                cell.fill = panos_fill
            elif ci == 4 and cs['prisma'] > cs['panos']:
                cell.fill = prisma_fill
        row += 1

    # Totals row
    row += 1
    ws.cell(row=row, column=1, value='TOTAL (Weighted)').font = Font(name='Calibri', size=11, bold=True)
    ws.cell(row=row, column=3, value=f'{result["panos_score"]:.0f}/100').font = Font(name='Calibri', size=11, bold=True, color=ORANGE)
    ws.cell(row=row, column=4, value=f'{result["prisma_score"]:.0f}/100').font = Font(name='Calibri', size=11, bold=True, color=TEAL)

    auto_width(ws)
    ws.column_dimensions['E'].width = 60


def _build_comparison_sheet(wb, result):
    ws = wb.create_sheet('Feature Comparison')

    row = 1
    cell = ws.cell(row=row, column=1, value='PAN-OS SD-WAN vs Prisma SD-WAN — Feature Comparison')
    cell.font = title_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 2

    headers = ['Feature', 'PAN-OS SD-WAN', 'Prisma SD-WAN', 'Advantage']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    row += 1

    for feat in result['feature_comparison']:
        advantage = feat.get('advantage', 'neutral')
        adv_label = 'PAN-OS' if advantage == 'panos' else ('Prisma' if advantage == 'prisma' else '-')

        for ci, val in enumerate([feat['feature'], feat['panos'], feat['prisma'], adv_label], 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border

            if ci == 2 and advantage == 'panos':
                cell.fill = panos_fill
            elif ci == 3 and advantage == 'prisma':
                cell.fill = prisma_fill
        row += 1

    auto_width(ws)
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 45


def _build_competitive_sheet(wb, result):
    competitive = result['competitive_displacement']
    ws = wb.create_sheet('Competitive Displacement')

    row = 1
    comp_label = competitive.get('competitor_label', 'Competitor')
    cell = ws.cell(row=row, column=1, value=f'Competitive Displacement: {comp_label}')
    cell.font = title_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 2

    # Key messages
    row = _write_section_header(ws, row, 'Key Messages', 2)
    for i, msg in enumerate(competitive.get('messages', []), 1):
        cell = ws.cell(row=row, column=1, value=f'{i}.')
        cell.font = Font(name='Calibri', size=10, bold=True, color=PURPLE)
        cell = ws.cell(row=row, column=2, value=msg)
        cell.font = data_font
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[row].height = 35
        row += 1

    row += 1

    # Objection handling
    if competitive.get('objections'):
        row = _write_section_header(ws, row, 'Objection Handling', 2)
        row = _write_table(ws, row, ['Objection', 'Response'],
                           [[obj, resp] for obj, resp in competitive['objections']])

    auto_width(ws)
    ws.column_dimensions['B'].width = 80


def _build_next_steps_sheet(wb, result):
    ws = wb.create_sheet('Next Steps')

    row = 1
    cell = ws.cell(row=row, column=1, value='Recommended Next Steps')
    cell.font = title_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 2

    for i, step in enumerate(result['next_steps'], 1):
        cell = ws.cell(row=row, column=1, value=i)
        cell.font = Font(name='Calibri', size=11, bold=True, color=BLUE)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        cell = ws.cell(row=row, column=2, value=step)
        cell.font = Font(name='Calibri', size=11)
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        ws.row_dimensions[row].height = 30
        row += 1

    auto_width(ws)
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 80
