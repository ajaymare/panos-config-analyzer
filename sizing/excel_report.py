"""Generate Excel report for sizing calculator results."""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from report.styles import (
    BLUE, WHITE, DARK_TEXT, BORDER_COLOR, GRAY,
    header_font, header_fill, header_align,
    data_font, data_align, thin_border, alt_fill,
    title_font, subtitle_font, kpi_label_font,
    kpi_value_font, style_kpi_cell, auto_width,
)

ORANGE = 'FA582D'
LIGHT_ORANGE = 'FDE8E0'
GREEN = '28A745'
LIGHT_GREEN = 'D4EDDA'

highlight_fill = PatternFill(start_color=LIGHT_ORANGE, end_color=LIGHT_ORANGE, fill_type='solid')
highlight_font = Font(name='Calibri', size=10, bold=True, color=ORANGE)


def _fmt(n):
    if isinstance(n, int):
        return f'{n:,}'
    return str(n)


def _write_section_header(ws, row, text, col_count):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name='Calibri', size=12, bold=True, color=BLUE)
    cell.alignment = Alignment(vertical='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    return row + 1


def _write_table(ws, start_row, headers, rows, highlight_model=None):
    """Write a table with headers and data rows. Returns next row."""
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for ri, row_data in enumerate(rows):
        r = start_row + 1 + ri
        is_highlighted = highlight_model and len(row_data) > 0 and row_data[0] == highlight_model
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = highlight_font if is_highlighted else data_font
            cell.alignment = data_align
            cell.border = thin_border
            if is_highlighted:
                cell.fill = highlight_fill
            elif ri % 2 == 1:
                cell.fill = alt_fill

    return start_row + 1 + len(rows) + 1


def generate_sizing_report(result, output_dir):
    """Generate the sizing recommendation Excel report.

    Args:
        result: dict from calculator.calculate_sizing()
        output_dir: directory to write the file

    Returns:
        str: path to generated Excel file
    """
    from .models import PA_MODELS, VM_MODELS, SECURITY_FEATURES

    wb = Workbook()
    hub = result['hub']
    branch = result['branch']
    summary = result['summary']
    licenses = result['licensing']
    security_features = result.get('security_features', {})
    hub_virtual = result.get('hub_virtual')
    vm_series = result.get('vm_series', False)

    # ========== Sheet 1: Sizing Recommendation ==========
    ws = wb.active
    ws.title = 'Sizing Recommendation'
    col_count = 6
    row = 1

    # Title
    cell = ws.cell(row=row, column=1, value='PAN-OS SD-WAN Sizing Recommendation')
    cell.font = Font(name='Calibri', size=16, bold=True, color=WHITE)
    cell.fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type='solid')
    cell.alignment = Alignment(vertical='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    row += 1

    cell = ws.cell(row=row, column=1, value=f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}')
    cell.font = subtitle_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    row += 2

    # KPI Row
    platform_label = 'Hardware + VM-Series' if vm_series else 'Hardware'
    hub_model_label = f'{hub["model"]}' + (f' / {hub_virtual["model"]}' if hub_virtual else '')
    style_kpi_cell(ws, row, 1, 'Hub Model', hub_model_label)
    style_kpi_cell(ws, row, 3, 'Branch Model', branch['model'])
    style_kpi_cell(ws, row, 5, 'Total Devices', summary['total_devices'])
    row += 3

    style_kpi_cell(ws, row, 1, 'Hub Sites', summary['num_hubs'])
    style_kpi_cell(ws, row, 3, 'Branch Sites', summary['num_branches'])
    style_kpi_cell(ws, row, 5, 'Platform', platform_label)
    row += 3

    hub_ha_label = 'Yes' if summary.get('hub_ha') else 'No'
    branch_ha_count = summary.get('branch_ha_count', 0)
    branch_ha_label = f'{branch_ha_count} of {summary["num_branches"]}' if branch_ha_count > 0 else 'No'
    style_kpi_cell(ws, row, 1, 'Hub HA', hub_ha_label)
    style_kpi_cell(ws, row, 3, 'Branch HA', branch_ha_label)
    style_kpi_cell(ws, row, 5, '', '')
    row += 3

    # --- Security Features ---
    row = _write_section_header(ws, row, 'Security Features', col_count)
    sec_rows = []
    for key, info in SECURITY_FEATURES.items():
        enabled = security_features.get(key, False)
        sec_rows.append([
            info['label'],
            info['description'],
            'Enabled' if enabled else 'Disabled',
            info['impact'].title(),
        ])
    row = _write_table(ws, row, ['Feature', 'Description', 'Status', 'Performance Impact'], sec_rows)

    # --- Hub Recommendation ---
    row = _write_section_header(ws, row, 'Hub Recommendation (Hardware)', col_count)
    hub_specs = [
        ['Model', hub['model']],
        ['Series', hub['specs'].get('series', '')],
        ['Platform', 'Hardware'],
        ['Description', hub['specs'].get('description', '')],
        ['Firewall Throughput', f'{_fmt(hub["specs"]["firewall_throughput"])} Mbps'],
        ['Threat Prevention Throughput', f'{_fmt(hub["specs"]["threat_throughput"])} Mbps'],
        ['SSL Decryption Throughput', f'{_fmt(hub["specs"].get("ssl_decrypt_throughput", 0))} Mbps'],
        ['IPSec VPN Throughput', f'{_fmt(hub["specs"]["ipsec_vpn_throughput"])} Mbps'],
        ['Max Concurrent Sessions', _fmt(hub['specs']['max_sessions'])],
        ['New Sessions/Second', _fmt(hub['specs']['new_sessions_per_sec'])],
        ['Max IPSec Tunnels', _fmt(hub['specs']['max_ipsec_tunnels'])],
        ['Max Security Rules', _fmt(hub['specs']['max_security_rules'])],
        ['Network Ports', hub['specs'].get('ports', '')],
        ['Form Factor', hub['specs'].get('form_factor', '')],
        ['Power Supply', hub['specs'].get('power_supply', '')],
        ['Device Count', hub['device_count']],
    ]
    row = _write_table(ws, row, ['Specification', 'Value'], hub_specs)

    # Hub Rationale
    row = _write_section_header(ws, row, 'Hub Sizing Rationale', col_count)
    for i, r in enumerate(hub['rationale']):
        cell = ws.cell(row=row, column=1, value=f'{i+1}. {r}')
        cell.font = data_font
        cell.alignment = data_align
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
        row += 1
    row += 1

    # --- Hub VM-Series Recommendation (if enabled) ---
    if hub_virtual:
        row = _write_section_header(ws, row, 'Hub Recommendation (VM-Series)', col_count)
        hub_vm_specs = [
            ['Model', hub_virtual['model']],
            ['Series', hub_virtual['specs'].get('series', '')],
            ['Platform', 'VM-Series'],
            ['Description', hub_virtual['specs'].get('description', '')],
            ['Firewall Throughput', f'{_fmt(hub_virtual["specs"]["firewall_throughput"])} Mbps'],
            ['Threat Prevention Throughput', f'{_fmt(hub_virtual["specs"]["threat_throughput"])} Mbps'],
            ['SSL Decryption Throughput', f'{_fmt(hub_virtual["specs"].get("ssl_decrypt_throughput", 0))} Mbps'],
            ['IPSec VPN Throughput', f'{_fmt(hub_virtual["specs"]["ipsec_vpn_throughput"])} Mbps'],
            ['Max Concurrent Sessions', _fmt(hub_virtual['specs']['max_sessions'])],
            ['New Sessions/Second', _fmt(hub_virtual['specs']['new_sessions_per_sec'])],
            ['Max IPSec Tunnels', _fmt(hub_virtual['specs']['max_ipsec_tunnels'])],
            ['Max Security Rules', _fmt(hub_virtual['specs']['max_security_rules'])],
            ['Form Factor', hub_virtual['specs'].get('form_factor', '')],
            ['Device Count', hub_virtual['device_count']],
        ]
        row = _write_table(ws, row, ['Specification', 'Value'], hub_vm_specs)

        row = _write_section_header(ws, row, 'Hub VM-Series Sizing Rationale', col_count)
        for i, r in enumerate(hub_virtual['rationale']):
            cell = ws.cell(row=row, column=1, value=f'{i+1}. {r}')
            cell.font = data_font
            cell.alignment = data_align
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
            row += 1
        row += 1

    # --- Branch Recommendation ---
    row = _write_section_header(ws, row, 'Branch Recommendation', col_count)
    branch_specs = [
        ['Model', branch['model']],
        ['Series', branch['specs'].get('series', '')],
        ['Platform', 'Hardware'],
        ['Description', branch['specs'].get('description', '')],
        ['Firewall Throughput', f'{_fmt(branch["specs"]["firewall_throughput"])} Mbps'],
        ['Threat Prevention Throughput', f'{_fmt(branch["specs"]["threat_throughput"])} Mbps'],
        ['SSL Decryption Throughput', f'{_fmt(branch["specs"].get("ssl_decrypt_throughput", 0))} Mbps'],
        ['IPSec VPN Throughput', f'{_fmt(branch["specs"]["ipsec_vpn_throughput"])} Mbps'],
        ['Max Concurrent Sessions', _fmt(branch['specs']['max_sessions'])],
        ['New Sessions/Second', _fmt(branch['specs']['new_sessions_per_sec'])],
        ['Max IPSec Tunnels', _fmt(branch['specs']['max_ipsec_tunnels'])],
        ['Max Security Rules', _fmt(branch['specs']['max_security_rules'])],
        ['Network Ports', branch['specs'].get('ports', '')],
        ['Form Factor', branch['specs'].get('form_factor', '')],
        ['Power Supply', branch['specs'].get('power_supply', '')],
        ['Device Count', branch['device_count']],
    ]
    row = _write_table(ws, row, ['Specification', 'Value'], branch_specs)

    # Branch Rationale
    row = _write_section_header(ws, row, 'Branch Sizing Rationale', col_count)
    for i, r in enumerate(branch['rationale']):
        cell = ws.cell(row=row, column=1, value=f'{i+1}. {r}')
        cell.font = data_font
        cell.alignment = data_align
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
        row += 1
    row += 1

    # --- Tunnel Calculation ---
    tunnel_calc = result.get('tunnel_calc', {})
    if tunnel_calc:
        row = _write_section_header(ws, row, 'IPSec Tunnel Calculation', col_count)
        bd = tunnel_calc.get('breakdown', {})
        tunnel_rows = [
            ['Hub -- Private ISPs', bd.get('hub', {}).get('private', ''), tunnel_calc.get('hub_private_tunnels', 0)],
            ['Hub -- Public ISPs', bd.get('hub', {}).get('public', ''), tunnel_calc.get('hub_public_tunnels', 0)],
            ['Hub -- Total per Hub', '', tunnel_calc.get('tunnels_per_hub', 0)],
            ['Branch -- Private ISPs', bd.get('branch', {}).get('private', ''), tunnel_calc.get('branch_private_tunnels', 0)],
            ['Branch -- Public ISPs', bd.get('branch', {}).get('public', ''), tunnel_calc.get('branch_public_tunnels', 0)],
            ['Branch -- Total per Branch', '', tunnel_calc.get('tunnels_per_branch', 0)],
        ]
        row = _write_table(ws, row, ['Component', 'Calculation', 'Tunnels'], tunnel_rows)

    # --- ISP Summary ---
    row = _write_section_header(ws, row, 'ISP Link Summary', col_count)
    hub_isps = hub.get('isps', {})
    branch_isps = branch.get('isps', {})
    isp_rows = [
        ['Hub', 'Public ISPs', hub_isps.get('public', 0)],
        ['Hub', 'Private ISPs', hub_isps.get('private', 0)],
        ['Hub', 'Total Links', hub_isps.get('total', 0)],
        ['Branch', 'Public ISPs', branch_isps.get('public', 0)],
        ['Branch', 'Private ISPs', branch_isps.get('private', 0)],
        ['Branch', 'Total Links', branch_isps.get('total', 0)],
    ]
    row = _write_table(ws, row, ['Role', 'ISP Type', 'Count'], isp_rows)

    # --- Licensing ---
    row = _write_section_header(ws, row, 'Licensing Recommendations', col_count)
    lic_rows = []
    for lic in licenses:
        lic_rows.append([
            lic['name'],
            lic.get('note', lic['description']),
            lic.get('applies_to', ''),
            'Required' if lic.get('required') else 'Recommended',
        ])
    row = _write_table(ws, row, ['License', 'Description', 'Applies To', 'Status'], lic_rows)

    # --- BOM ---
    row = _write_section_header(ws, row, 'Bill of Materials', col_count)
    hub_ha_str = 'Yes' if summary.get('hub_ha') else 'No'
    bha = summary.get('branch_ha_count', 0)
    branch_ha_str = f'{bha} sites' if bha > 0 else 'No'
    bom_rows = [
        ['Hub', hub['model'], 'Hardware', summary['num_hubs'], hub_ha_str, hub['device_count']],
    ]
    if hub_virtual:
        bom_rows.append(['Hub (Cloud)', hub_virtual['model'], 'VM-Series', summary['num_hubs'], hub_ha_str, hub_virtual['device_count']])
    bom_rows.append(['Branch', branch['model'], 'Hardware', summary['num_branches'], branch_ha_str, branch['device_count']])
    bom_rows.append(['Total', '', '', '', '', summary['total_devices']])
    row = _write_table(ws, row, ['Role', 'Model', 'Platform', 'Sites', 'HA', 'Devices'], bom_rows)

    auto_width(ws)

    # ========== Sheet 2+: Model Comparison ==========
    # Always show hardware; add VM-Series sheet when enabled
    catalogs = [('Hardware Model Comparison', PA_MODELS)]
    if vm_series:
        catalogs.append(('VM-Series Comparison', VM_MODELS))

    headers = [
        'Model', 'Series', 'FW Throughput (Mbps)', 'TP Throughput (Mbps)',
        'SSL Decrypt (Mbps)', 'IPSec Throughput (Mbps)', 'Max Sessions',
        'New Sess/Sec', 'Max IPSec Tunnels', 'Max Rules', 'Form Factor',
        'Ports', 'Role', 'Description',
    ]

    for sheet_title, models in catalogs:
        ws2 = wb.create_sheet(sheet_title)

        # Title
        cell = ws2.cell(row=1, column=1, value=f'{sheet_title}')
        cell.font = Font(name='Calibri', size=14, bold=True, color=BLUE)
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

        cell = ws2.cell(row=2, column=1,
                        value=f'Hub recommendation: {hub["model"]}  |  Branch recommendation: {branch["model"]}')
        cell.font = subtitle_font
        ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

        model_rows = []
        for model_name, specs in models.items():
            model_rows.append([
                model_name,
                specs.get('series', ''),
                specs['firewall_throughput'],
                specs['threat_throughput'],
                specs.get('ssl_decrypt_throughput', 0),
                specs['ipsec_vpn_throughput'],
                specs['max_sessions'],
                specs['new_sessions_per_sec'],
                specs['max_ipsec_tunnels'],
                specs['max_security_rules'],
                specs.get('form_factor', ''),
                specs.get('ports', ''),
                specs.get('recommended_role', '').title(),
                specs.get('description', ''),
            ])

        start = 4
        for ci, h in enumerate(headers, 1):
            cell = ws2.cell(row=start, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for ri, row_data in enumerate(model_rows):
            r = start + 1 + ri
            model_name = row_data[0]
            is_hub = model_name == hub['model'] or (hub_virtual and model_name == hub_virtual['model'])
            is_branch = model_name == branch['model']
            for ci, val in enumerate(row_data, 1):
                cell = ws2.cell(row=r, column=ci, value=val)
                cell.border = thin_border
                cell.alignment = data_align
                if is_hub or is_branch:
                    cell.fill = highlight_fill
                    cell.font = highlight_font
                else:
                    cell.font = data_font
                    if ri % 2 == 1:
                        cell.fill = alt_fill

        auto_width(ws2, max_width=40)

    # Save
    filename = f'Sizing_Recommendation_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath
