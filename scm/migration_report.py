"""Generate SCM Migration Report Excel — shows conversion status for each feature.

Produces an Excel workbook with:
  - Migration Summary sheet: per-feature conversion status
  - Converted Details: items that were fully converted
  - Partially Converted: items with some fields not mappable
  - Not Supported: features with no SCM equivalent
"""
from __future__ import annotations

import io
import os
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from parsers.base import FeatureResult
from scm.mapper import get_mappers, get_mapper_for_feature, map_results

# Display names for SCM resource types
_ROLE_DISPLAY_NAMES = {
    'sdwan_interface_profiles': 'SD-WAN Interface Profiles',
    'sdwan_path_quality_profiles': 'Path Quality Profiles',
    'sdwan_traffic_distribution_profiles': 'Traffic Distribution Profiles',
    'sdwan_policies': 'SD-WAN Policies (App-ID Steering)',
    'security_rules': 'Security Rules',
    'nat_rules': 'NAT Rules',
    'ike_gateways': 'IKE Gateways',
    'ipsec_tunnels': 'IPSec Tunnels',
    'bgp_routing_profiles': 'BGP Routing Profiles',
    'interfaces': 'Interfaces',
    'zones': 'Zones',
    'custom_applications': 'Custom Applications',
    'sdwan_vpn_clusters': 'SD-WAN VPN Clusters',
    'sdwan_devices': 'SD-WAN Devices',
    'sdwan_bgp_policies': 'SD-WAN BGP Policies',
}

# ---------------------------------------------------------------------------
# Feature categories (mirrors excel_generator.py)
# ---------------------------------------------------------------------------

FEATURE_CATEGORIES = {
    'SD-WAN Core': [
        'SD-WAN Interface Profiles', 'App-ID Steering', 'Path Quality Metrics',
        'Bandwidth Monitoring', 'Probe Idle Time', 'Failback Hold Time',
    ],
    'Traffic Optimization': [
        'Link Remediation (FEC)', 'Packet Duplication',
    ],
    'VPN & Topology': [
        'VPN Automation', 'Topology Configured',
        'Hub Capacity', 'Prisma Access Hub',
        'Sub-Second Failover', 'Tunnel Monitor',
    ],
    'Routing': [
        'Dynamic Routing', 'BGP AS Control', 'BGP Private AS',
        'BGP Timer Profile', 'BGP Security Rule',
        'BGP Routing Profiles', 'BGP Dampening',
        'IPv6 Support', 'Multi-VR Support',
        'Multicast Support', 'BFD Configuration',
        'Advance Routing',
    ],
    'Security & NAT': [
        'SD-WAN Security Rules', 'SD-WAN NAT Policies',
    ],
    'Monitoring & Reporting': [
        'ADEM Integration', 'SD-WAN Reporting',
        'Log Collection', 'Device Telemetry',
        'Monitor Profiles',
    ],
    'Network Infrastructure': [
        'Sub/Agg Interfaces', 'Custom Applications',
        'Template/Stack Mapping', 'Upstream NAT',
        'ZTP Support',
    ],
}

CAT_COLORS = {
    'SD-WAN Core': '1B4F72',
    'Traffic Optimization': '884EA0',
    'VPN & Topology': '6C3483',
    'Routing': '1E8449',
    'Security & NAT': 'C0392B',
    'Monitoring & Reporting': 'B9770E',
    'Network Infrastructure': '2E86C1',
}

# Features that have a direct SCM mapper (feature_name → scm_resource_name)
# Built from mapper registry
_SCM_SUPPORTED_FEATURES: dict[str, str] = {}
for _m in get_mappers():
    _SCM_SUPPORTED_FEATURES.setdefault(_m.feature_name, _m.scm_resource_name)

# Features that are sub-features / boolean flags — partially supported via parent
_PARTIAL_FEATURES = {
    'Bandwidth Monitoring': 'Configured via SD-WAN Interface Profiles (maximum_upload/download)',
    'Probe Idle Time': 'Configured via SD-WAN Interface Profiles (probe_idle_time)',
    'Failback Hold Time': 'Configured via SD-WAN Interface Profiles (failback_hold_time)',
    'Link Remediation (FEC)': 'Configured via Traffic Distribution Profiles (error_correction)',
    'Packet Duplication': 'Configured via Traffic Distribution Profiles (packet_duplication)',
    'Tunnel Monitor': 'Configured via IPSec Tunnels (tunnel_monitor)',
    'BGP AS Control': 'Configured via SD-WAN Devices (bgp/as_number)',
    'BGP Private AS': 'Configured via SD-WAN Devices (bgp/remove_private_as)',
    'BGP Timer Profile': 'Configured via BGP Routing Profiles (graceful_restart)',
    'BGP Dampening': 'Configured via BGP Routing Profiles',
    'Topology Configured': 'Configured via SD-WAN VPN Clusters (type field)',
    'Hub Capacity': 'Configured via SD-WAN VPN Clusters (hubs/branches)',
    'Sub-Second Failover': 'Configured via SD-WAN VPN Clusters (dia_vpn_failover)',
    'Multi-VR Support': 'Configured via SD-WAN Devices (multi_vr_support)',
    'BGP Security Rule': 'Configured via SD-WAN BGP Policies (rules)',
}

# Features not supported by SCM API
_NOT_SUPPORTED_REASONS = {
    'Prisma Access Hub': 'Prisma Access hub configured via SCM Prisma Access settings',
    'BGP Routing Profiles': 'SCM manages BGP profiles via SD-WAN routing API',
    'IPv6 Support': 'IPv6 routing configured directly in SCM network settings',
    'Multicast Support': 'Multicast not yet available via SCM API',
    'BFD Configuration': 'BFD configured in SCM routing settings',
    'Advance Routing': 'Logical routers managed in SCM advanced routing',
    'ADEM Integration': 'ADEM configured in SCM Autonomous DEM settings',
    'SD-WAN Reporting': 'Reporting is native to SCM — no migration needed',
    'Log Collection': 'Log forwarding configured in SCM log settings',
    'Device Telemetry': 'Telemetry enabled in SCM device settings',
    'Monitor Profiles': 'Monitor profiles configured in SCM network monitoring',
    'Template/Stack Mapping': 'Templates not applicable — SCM uses folder-based management',
    'Upstream NAT': 'Upstream NAT configured on SCM interface settings',
    'ZTP Support': 'ZTP is native to SCM — devices auto-onboard',
}

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
_HEADER_FILL = PatternFill('solid', fgColor='1a2a44')
_CAT_FONT = Font(bold=True, color='FFFFFF', size=10)
_CONVERTED_FILL = PatternFill('solid', fgColor='D5F5E3')
_PARTIAL_FILL = PatternFill('solid', fgColor='FEF9E7')
_NOT_SUPPORTED_FILL = PatternFill('solid', fgColor='FADBD8')
_IGNORED_FILL = PatternFill('solid', fgColor='E8DAEF')
_CONVERTED_FONT = Font(color='1E8449', bold=True)
_PARTIAL_FONT = Font(color='B9770E', bold=True)
_NOT_SUPPORTED_FONT = Font(color='C0392B', bold=True)
_IGNORED_FONT = Font(color='6C3483', bold=True)
_THIN_BORDER = Border(
    left=Side(style='thin', color='D4DBE6'),
    right=Side(style='thin', color='D4DBE6'),
    top=Side(style='thin', color='D4DBE6'),
    bottom=Side(style='thin', color='D4DBE6'),
)


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value) if cell.value else ''
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 60)


# ---------------------------------------------------------------------------
# Classification logic
# ---------------------------------------------------------------------------

def _classify_feature(
    feature_name: str,
    enabled: bool,
    selected_features: list[str] | None,
) -> tuple[str, str, str]:
    """Classify a feature's migration status.

    Returns: (status, scm_reference, notes)
      status: 'Fully Converted', 'Partially Converted', 'Not Supported by SCM', 'Ignored (User Skipped)', 'Not Configured'
    """
    # Check if feature is configured at all
    if not enabled:
        return 'Not Configured', '', 'Feature not enabled in PAN-OS config'

    # Check if feature has a direct mapper
    if feature_name in _SCM_SUPPORTED_FEATURES:
        resource = _SCM_SUPPORTED_FEATURES[feature_name]
        mapper = get_mapper_for_feature(feature_name)
        endpoint = mapper[0].scm_endpoint if mapper else ''

        # Check if user excluded it
        if selected_features is not None and resource not in selected_features:
            return 'Ignored (User Skipped)', endpoint, f'User chose not to migrate this feature'

        return 'Fully Converted', endpoint, f'Ansible role: {resource}'

    # Check if it's a partial/sub-feature
    if feature_name in _PARTIAL_FEATURES:
        parent_note = _PARTIAL_FEATURES[feature_name]
        return 'Partially Converted', '', parent_note

    # Not supported
    reason = _NOT_SUPPORTED_REASONS.get(feature_name, 'No SCM API endpoint available')
    return 'Not Supported by SCM', '', reason


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

def generate_migration_report(
    results: list[FeatureResult],
    selected_features: list[str] | None = None,
    mapped: dict[str, dict] | None = None,
) -> bytes:
    """Generate migration report Excel as bytes.

    Args:
        results: All FeatureResult objects from parsing.
        selected_features: Which features the user selected for migration.
        mapped: Output from map_results() (already filtered).

    Returns:
        Excel file content as bytes.
    """
    if mapped is None:
        mapped = map_results(results)

    # Build a lookup of enabled features
    enabled_features: dict[str, bool] = {}
    for r in results:
        if r.feature_name not in enabled_features:
            enabled_features[r.feature_name] = r.enabled
        elif r.enabled:
            enabled_features[r.feature_name] = True

    wb = Workbook()

    # ── Sheet 1: Migration Summary ─────────────────────────────
    ws = wb.active
    ws.title = 'Migration Summary'

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = 'PAN-OS to SCM Migration Report'
    ws['A1'].font = Font(bold=True, size=14, color='1a2a44')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Counters
    counts = {'Fully Converted': 0, 'Partially Converted': 0,
              'Not Supported by SCM': 0, 'Ignored (User Skipped)': 0,
              'Not Configured': 0}

    # Headers
    headers = ['Category', 'Feature', 'PAN-OS Status', 'Migration Status',
               'SCM API Reference', 'Notes']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = _THIN_BORDER

    row_num = 4
    all_rows = []  # for detail sheets

    for category, features in FEATURE_CATEGORIES.items():
        # Category header row
        cat_color = CAT_COLORS.get(category, '333333')
        cat_fill = PatternFill('solid', fgColor=cat_color)
        for col in range(1, 7):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = cat_fill
            cell.font = _CAT_FONT
            cell.border = _THIN_BORDER
        ws.cell(row=row_num, column=1, value=category)
        row_num += 1

        for feat in features:
            enabled = enabled_features.get(feat, False)
            status, scm_ref, notes = _classify_feature(feat, enabled, selected_features)
            counts[status] = counts.get(status, 0) + 1

            panos_status = 'Enabled' if enabled else 'Not Configured'

            ws.cell(row=row_num, column=1, value=category).border = _THIN_BORDER
            ws.cell(row=row_num, column=2, value=feat).border = _THIN_BORDER
            ws.cell(row=row_num, column=3, value=panos_status).border = _THIN_BORDER

            status_cell = ws.cell(row=row_num, column=4, value=status)
            status_cell.border = _THIN_BORDER
            if status == 'Fully Converted':
                status_cell.fill = _CONVERTED_FILL
                status_cell.font = _CONVERTED_FONT
            elif status == 'Partially Converted':
                status_cell.fill = _PARTIAL_FILL
                status_cell.font = _PARTIAL_FONT
            elif status == 'Not Supported by SCM':
                status_cell.fill = _NOT_SUPPORTED_FILL
                status_cell.font = _NOT_SUPPORTED_FONT
            elif status == 'Ignored (User Skipped)':
                status_cell.fill = _IGNORED_FILL
                status_cell.font = _IGNORED_FONT

            ws.cell(row=row_num, column=5, value=scm_ref).border = _THIN_BORDER
            ws.cell(row=row_num, column=6, value=notes).border = _THIN_BORDER

            all_rows.append({
                'category': category, 'feature': feat,
                'panos_status': panos_status, 'status': status,
                'scm_ref': scm_ref, 'notes': notes,
            })

            row_num += 1

    # Summary counters at top
    row_num += 1
    ws.cell(row=row_num, column=1, value='Migration Summary').font = Font(bold=True, size=12)
    row_num += 1
    status_styles = {
        'Fully Converted': (_CONVERTED_FILL, _CONVERTED_FONT),
        'Partially Converted': (_PARTIAL_FILL, _PARTIAL_FONT),
        'Not Supported by SCM': (_NOT_SUPPORTED_FILL, _NOT_SUPPORTED_FONT),
        'Ignored (User Skipped)': (_IGNORED_FILL, _IGNORED_FONT),
        'Not Configured': (PatternFill('solid', fgColor='F2F3F4'), Font(color='666666')),
    }
    for label, count in counts.items():
        ws.cell(row=row_num, column=1, value=label).border = _THIN_BORDER
        fill, font = status_styles.get(label, (None, None))
        c = ws.cell(row=row_num, column=1)
        if fill:
            c.fill = fill
        if font:
            c.font = font
        ws.cell(row=row_num, column=2, value=count).border = _THIN_BORDER
        ws.cell(row=row_num, column=2).font = Font(bold=True)
        row_num += 1

    _auto_width(ws)

    # ── Sheet 2: Converted Details ─────────────────────────────
    ws2 = wb.create_sheet('Converted Details')
    ws2.merge_cells('A1:E1')
    ws2['A1'] = 'Fully Converted Features — Ready for SCM Deployment'
    ws2['A1'].font = Font(bold=True, size=13, color='1E8449')

    detail_headers = ['Feature', 'SCM Resource', 'SCM Endpoint', 'Items Converted', 'SCM Folder']
    for col, h in enumerate(detail_headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = PatternFill('solid', fgColor='1E8449')
        cell.border = _THIN_BORDER

    drow = 4
    for resource_name, info in sorted(mapped.items()):
        display = _ROLE_DISPLAY_NAMES.get(resource_name, resource_name)
        ws2.cell(row=drow, column=1, value=display).border = _THIN_BORDER
        ws2.cell(row=drow, column=2, value=resource_name).border = _THIN_BORDER
        ws2.cell(row=drow, column=3, value=info['endpoint']).border = _THIN_BORDER
        ws2.cell(row=drow, column=4, value=len(info['payloads'])).border = _THIN_BORDER
        ws2.cell(row=drow, column=4).alignment = Alignment(horizontal='center')
        ws2.cell(row=drow, column=5, value=info['folder']).border = _THIN_BORDER

        # Fill with converted style
        for col in range(1, 6):
            ws2.cell(row=drow, column=col).fill = _CONVERTED_FILL
        drow += 1

        # List individual items
        for p in info['payloads']:
            name = p.get('name', 'unnamed')
            ws2.cell(row=drow, column=2, value=f'  {name}').border = _THIN_BORDER
            # Show key fields
            fields = [f'{k}={v}' for k, v in p.items() if k != 'name']
            ws2.cell(row=drow, column=3, value=', '.join(fields[:5])).border = _THIN_BORDER
            drow += 1

    _auto_width(ws2)

    # ── Sheet 3: Partially Converted ───────────────────────────
    ws3 = wb.create_sheet('Partially Converted')
    ws3.merge_cells('A1:D1')
    ws3['A1'] = 'Partially Converted Features — Included via Parent Feature'
    ws3['A1'].font = Font(bold=True, size=13, color='B9770E')

    partial_headers = ['Feature', 'Category', 'How It Maps to SCM', 'Action Required']
    for col, h in enumerate(partial_headers, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = PatternFill('solid', fgColor='B9770E')
        cell.border = _THIN_BORDER

    prow = 4
    for r in all_rows:
        if r['status'] != 'Partially Converted':
            continue
        ws3.cell(row=prow, column=1, value=r['feature']).border = _THIN_BORDER
        ws3.cell(row=prow, column=2, value=r['category']).border = _THIN_BORDER
        ws3.cell(row=prow, column=3, value=r['notes']).border = _THIN_BORDER
        ws3.cell(row=prow, column=4, value='Verify in parent playbook vars').border = _THIN_BORDER
        for col in range(1, 5):
            ws3.cell(row=prow, column=col).fill = _PARTIAL_FILL
        prow += 1

    if prow == 4:
        ws3.cell(row=4, column=1, value='No partially converted features')

    _auto_width(ws3)

    # ── Sheet 4: Not Supported ─────────────────────────────────
    ws4 = wb.create_sheet('Not Supported')
    ws4.merge_cells('A1:D1')
    ws4['A1'] = 'Features Not Supported via SCM API — Manual Configuration Required'
    ws4['A1'].font = Font(bold=True, size=13, color='C0392B')

    ns_headers = ['Feature', 'Category', 'Reason', 'Recommendation']
    for col, h in enumerate(ns_headers, 1):
        cell = ws4.cell(row=3, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = PatternFill('solid', fgColor='C0392B')
        cell.border = _THIN_BORDER

    nrow = 4
    for r in all_rows:
        if r['status'] != 'Not Supported by SCM':
            continue
        ws4.cell(row=nrow, column=1, value=r['feature']).border = _THIN_BORDER
        ws4.cell(row=nrow, column=2, value=r['category']).border = _THIN_BORDER
        ws4.cell(row=nrow, column=3, value=r['notes']).border = _THIN_BORDER
        ws4.cell(row=nrow, column=4, value='Configure manually in SCM UI').border = _THIN_BORDER
        for col in range(1, 5):
            ws4.cell(row=nrow, column=col).fill = _NOT_SUPPORTED_FILL
        nrow += 1

    if nrow == 4:
        ws4.cell(row=4, column=1, value='All configured features are supported')

    _auto_width(ws4)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
