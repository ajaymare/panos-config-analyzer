"""Excel cell styles and formatting constants."""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle

# Colors
BLUE = '1A2A44'
LIGHT_BLUE = 'D6E4F0'
GREEN = '28A745'
LIGHT_GREEN = 'D4EDDA'
RED = 'DC3545'
LIGHT_RED = 'F8D7DA'
GRAY = 'F2F2F2'
WHITE = 'FFFFFF'
DARK_TEXT = '1E2A3A'
BORDER_COLOR = 'D4DBE6'

# Borders
thin_border = Border(
    left=Side(style='thin', color=BORDER_COLOR),
    right=Side(style='thin', color=BORDER_COLOR),
    top=Side(style='thin', color=BORDER_COLOR),
    bottom=Side(style='thin', color=BORDER_COLOR),
)

# Header style
header_font = Font(name='Calibri', size=11, bold=True, color=WHITE)
header_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Data styles
data_font = Font(name='Calibri', size=10, color=DARK_TEXT)
data_align = Alignment(vertical='center', wrap_text=True)

# Status fills
enabled_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type='solid')
enabled_font = Font(name='Calibri', size=10, bold=True, color=GREEN)
disabled_fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type='solid')
disabled_font = Font(name='Calibri', size=10, bold=True, color=RED)

# Alternating row fill
alt_fill = PatternFill(start_color=GRAY, end_color=GRAY, fill_type='solid')

# Title style
title_font = Font(name='Calibri', size=14, bold=True, color=BLUE)
subtitle_font = Font(name='Calibri', size=11, color='6B7A8D')


def style_header_row(ws, row_num, col_count):
    """Apply header styling to a row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def style_data_cell(cell, row_idx=0):
    """Apply data styling to a cell."""
    cell.font = data_font
    cell.alignment = data_align
    cell.border = thin_border
    if row_idx % 2 == 1:
        cell.fill = alt_fill


def style_status_cell(cell, enabled):
    """Apply enabled/disabled styling."""
    cell.border = thin_border
    if enabled:
        cell.fill = enabled_fill
        cell.font = enabled_font
        cell.value = 'Enabled'
    else:
        cell.fill = disabled_fill
        cell.font = disabled_font
        cell.value = 'Disabled'
    cell.alignment = Alignment(horizontal='center', vertical='center')


# KPI styles
kpi_label_font = Font(name='Calibri', size=10, bold=True, color='6B7A8D')
kpi_label_align = Alignment(horizontal='center', vertical='center')
kpi_value_font = Font(name='Calibri', size=18, bold=True, color=DARK_TEXT)
kpi_value_align = Alignment(horizontal='center', vertical='center')

# Level fills
level_fills = {
    'Full': PatternFill(start_color='1E8449', end_color='1E8449', fill_type='solid'),
    'Advanced': PatternFill(start_color='B9770E', end_color='B9770E', fill_type='solid'),
    'Basic': PatternFill(start_color='2E86C1', end_color='2E86C1', fill_type='solid'),
}
level_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')


def progress_bar(percent, width=10):
    """Return a text-based progress bar: ████████░░ (80%)."""
    filled = round(percent / 100 * width)
    empty = width - filled
    return '\u2588' * filled + '\u2591' * empty


def style_kpi_cell(ws, row, col, label, value, value_font=None, value_fill=None, merge_cols=2):
    """Write a KPI label+value pair into two rows at (row, col)."""
    # Label
    cell_label = ws.cell(row=row, column=col, value=label)
    cell_label.font = kpi_label_font
    cell_label.alignment = kpi_label_align
    cell_label.border = thin_border
    if merge_cols > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + merge_cols - 1)

    # Value
    cell_value = ws.cell(row=row + 1, column=col, value=value)
    cell_value.font = value_font or kpi_value_font
    cell_value.alignment = kpi_value_align
    cell_value.border = thin_border
    if value_fill:
        cell_value.fill = value_fill
    if merge_cols > 1:
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + merge_cols - 1)


def auto_width(ws, min_width=10, max_width=50):
    """Auto-adjust column widths based on content."""
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        if col_letter:
            adjusted = min(max(max_len + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = adjusted
