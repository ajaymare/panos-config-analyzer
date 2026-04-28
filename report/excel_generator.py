"""Generate Excel report from parser results."""
import os
import uuid
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from . import styles
import config as app_config
REPORT_DIR = app_config.REPORT_DIR

# Category grouping for quick reference
FEATURE_CATEGORIES = {
    'SD-WAN Core': [
        'SD-WAN Interface Profiles', 'App-ID Steering', 'Path Quality Metrics',
        'Bandwidth Monitoring', 'Probe Idle Time', 'Failback Hold Time',
    ],
    'Traffic Optimization': [
        'Link Remediation (FEC)', 'Packet Duplication',
    ],
    'VPN & Topology': [
        'VPN Automation', 'Topologies Supported',
        'Hub Capacity', 'Prisma Access Hub',
        'Sub-Second Failover', 'Tunnel Monitor',
    ],
    'Routing': [
        'Dynamic Routing', 'BGP AS Control', 'BGP Private AS',
        'BGP Timer Profile', 'BGP Security Rule',
        'BGP Routing Profiles', 'BGP Dampening',
        'IPv6 Support', 'Multi-VR Support',
        'Multicast Support', 'BFD Configuration',
        'Advance Routing',
    ],
    'Security & NAT': [
        'SD-WAN Security Rules', 'SD-WAN NAT Policies',
    ],
    'Monitoring & Reporting': [
        'ADEM Integration', 'SD-WAN Reporting',
        'Log Collection', 'Device Telemetry',
        'Monitor Profiles',
    ],
    'Network Infrastructure': [
        'Sub/Agg Interfaces', 'Custom Applications',
        'Template/Stack Mapping', 'Upstream NAT',
        'ZTP Support',
    ],
}

# Category colors
CAT_COLORS = {
    'SD-WAN Core': '1B4F72',
    'Traffic Optimization': '884EA0',
    'VPN & Topology': '6C3483',
    'Routing': '1E8449',
    'Security & NAT': 'C0392B',
    'Monitoring & Reporting': 'B9770E',
    'Network Infrastructure': '2E86C1',
}


def _add_executive_summary(wb, scored_list, is_first_sheet=True):
    """Add an Executive Summary sheet with deployment scoring and gap analysis.

    Args:
        wb: openpyxl Workbook
        scored_list: list of dicts from scorer.score_configs() or [scorer.score_config()]
        is_first_sheet: if True, use the active sheet; otherwise create new
    """
    if is_first_sheet:
        ws = wb.active
        ws.title = 'Executive Summary'
    else:
        ws = wb.create_sheet(title='Executive Summary', index=0)

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'].value = 'SD-WAN Deployment Executive Summary'
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color=styles.BLUE)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:F2')
    ws['A2'].value = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = styles.subtitle_font
    ws['A2'].alignment = Alignment(horizontal='center')

    row = 4

    # Per-config scoring
    for s in scored_list:
        sc = s['scoring']
        name = s.get('filename', 'Config')
        cfg_type = s.get('config_type', 'unknown')

        # Config header
        header_fill = PatternFill(start_color='1a2a44', end_color='1a2a44', fill_type='solid')
        header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = styles.thin_border
        ws.cell(row=row, column=1, value=f'{name} ({cfg_type.upper()})')
        row += 1

        # Software version info
        versions = s.get('versions')
        if versions:
            version_parts = []
            if versions.get('panos_version'):
                version_parts.append(f'PAN-OS {versions["panos_version"]}')
            if versions.get('sdwan_version'):
                version_parts.append(f'SD-WAN Plugin {versions["sdwan_version"]}')
            if version_parts:
                ws.cell(row=row, column=1, value='Software Version')
                ws.cell(row=row, column=1).font = Font(name='Calibri', size=11, bold=True)
                ws.cell(row=row, column=2, value=' | '.join(version_parts))
                ws.cell(row=row, column=2).font = Font(name='Calibri', size=11, color='0066CC')
                ws.cell(row=row, column=2).border = styles.thin_border
                row += 1

        # Devices/Sources info — extract unique sources from results
        results = s.get('results', [])
        if results:
            sources = []
            for r in results:
                if r.source and r.source not in sources and r.enabled:
                    sources.append(r.source)
            if sources:
                ws.cell(row=row, column=1, value='Devices / Sources')
                ws.cell(row=row, column=1).font = Font(name='Calibri', size=11, bold=True)
                ws.cell(row=row, column=2, value=', '.join(sources))
                ws.cell(row=row, column=2).font = Font(name='Calibri', size=11, color='1a2a44')
                ws.cell(row=row, column=2).border = styles.thin_border
                row += 1

        # Score and level
        level_colors = {'Full': '1E8449', 'Advanced': 'B9770E', 'Basic': '2E86C1'}
        level_color = level_colors.get(sc['level'], '2E86C1')
        level_fill = PatternFill(start_color=level_color, end_color=level_color, fill_type='solid')

        ws.cell(row=row, column=1, value='Deployment Maturity')
        ws.cell(row=row, column=1).font = Font(name='Calibri', size=11, bold=True)
        cell = ws.cell(row=row, column=2, value=sc['level'])
        cell.fill = level_fill
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
        cell.border = styles.thin_border

        ws.cell(row=row, column=3, value='Score')
        ws.cell(row=row, column=3).font = Font(name='Calibri', size=11, bold=True)
        ws.cell(row=row, column=4, value=f'{sc["score"]}/{sc["total"]} ({sc["percent"]}%)')
        ws.cell(row=row, column=4).font = Font(name='Calibri', size=11, bold=True, color=level_color)
        row += 2

        # Category breakdown with coverage
        cat_headers = ['Category', 'Enabled', 'Total', 'Coverage']
        for col, h in enumerate(cat_headers, 1):
            ws.cell(row=row, column=col, value=h)
        styles.style_header_row(ws, row, len(cat_headers))
        row += 1

        for cat_name, cat_data in sc['category_scores'].items():
            cat_color = CAT_COLORS.get(cat_name, '2E86C1')
            ws.cell(row=row, column=1, value=cat_name)
            styles.style_data_cell(ws.cell(row=row, column=1), row)

            ws.cell(row=row, column=2, value=cat_data['enabled'])
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2).border = styles.thin_border
            if cat_data['enabled'] == cat_data['total']:
                ws.cell(row=row, column=2).font = Font(name='Calibri', size=11, bold=True, color='1E8449')
            elif cat_data['enabled'] > 0:
                ws.cell(row=row, column=2).font = Font(name='Calibri', size=11, color='B9770E')
            else:
                ws.cell(row=row, column=2).font = Font(name='Calibri', size=11, color='C0392B')

            ws.cell(row=row, column=3, value=cat_data['total'])
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3).border = styles.thin_border

            ws.cell(row=row, column=4, value=f'{cat_data["percent"]}%')
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=4).border = styles.thin_border
            row += 1

        row += 1

        # Feature Details — show what's configured per device/source
        results = s.get('results', [])
        if results:
            # Group results by feature
            feat_groups = {}
            for r in results:
                if r.feature_name not in feat_groups:
                    feat_groups[r.feature_name] = []
                feat_groups[r.feature_name].append(r)

            detail_headers = ['Device', 'Feature', 'Status', 'Enabled Count']
            for col, h in enumerate(detail_headers, 1):
                ws.cell(row=row, column=col, value=h)
            styles.style_header_row(ws, row, len(detail_headers))
            row += 1

            for cat_name, features in FEATURE_CATEGORIES.items():
                cat_color = CAT_COLORS.get(cat_name, '2E86C1')
                cat_fill = PatternFill(start_color=cat_color, end_color=cat_color, fill_type='solid')
                cat_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                for col in range(1, len(detail_headers) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = cat_fill
                    cell.font = cat_font
                    cell.border = styles.thin_border
                ws.cell(row=row, column=1, value=cat_name)
                row += 1

                for feat_name in features:
                    rlist = feat_groups.get(feat_name, [])
                    enabled_count = sum(1 for r in rlist if r.enabled)

                    ws.cell(row=row, column=1, value=name)
                    styles.style_data_cell(ws.cell(row=row, column=1), row)
                    ws.cell(row=row, column=2, value=feat_name)
                    styles.style_data_cell(ws.cell(row=row, column=2), row)
                    styles.style_status_cell(ws.cell(row=row, column=3), enabled_count > 0)

                    ws.cell(row=row, column=4, value=enabled_count)
                    ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
                    ws.cell(row=row, column=4).border = styles.thin_border
                    if enabled_count > 0:
                        ws.cell(row=row, column=4).font = Font(name='Calibri', size=11, bold=True, color='1E8449')
                    else:
                        ws.cell(row=row, column=4).font = Font(name='Calibri', size=11, color='C0392B')
                    row += 1

            row += 1

        # Panorama-Managed features
        pan_managed = sc.get('panorama_managed_features', [])
        if pan_managed:
            ws.cell(row=row, column=1, value='Panorama-Managed Features')
            ws.cell(row=row, column=1).font = Font(name='Calibri', size=11, bold=True, color='B9770E')
            row += 1
            for feat in pan_managed:
                ws.cell(row=row, column=1, value=f'  {feat}')
                ws.cell(row=row, column=1).font = Font(name='Calibri', size=11, color='B9770E')
                ws.cell(row=row, column=1).border = styles.thin_border
                ws.cell(row=row, column=2, value='Configured via Panorama')
                ws.cell(row=row, column=2).font = Font(name='Calibri', size=11, color='B9770E')
                ws.cell(row=row, column=2).border = styles.thin_border
                row += 1
            row += 1

        # Missing features / recommendations
        if sc['missing_features']:
            ws.cell(row=row, column=1, value='Recommendations')
            ws.cell(row=row, column=1).font = Font(name='Calibri', size=11, bold=True, color='C0392B')
            row += 1
            rec_headers = ['Missing Feature', 'Recommendation']
            for col, h in enumerate(rec_headers, 1):
                ws.cell(row=row, column=col, value=h)
            styles.style_header_row(ws, row, len(rec_headers))
            row += 1

            for feat, rec in zip(sc['missing_features'], sc['recommendations']):
                ws.cell(row=row, column=1, value=feat)
                styles.style_data_cell(ws.cell(row=row, column=1), row)
                ws.cell(row=row, column=2, value=rec)
                styles.style_data_cell(ws.cell(row=row, column=2), row)
                row += 1
        elif not pan_managed:
            ws.cell(row=row, column=1, value='All SD-WAN features are configured.')
            ws.cell(row=row, column=1).font = Font(name='Calibri', size=11, bold=True, color='1E8449')
            row += 1

        row += 2  # Space between configs

    styles.auto_width(ws, min_width=14, max_width=60)
    ws.freeze_panes = 'A4'


def generate(results: list, config_type: str = 'unknown', versions: dict = None, output_dir: str = None) -> str:
    wb = Workbook()

    # ── Executive Summary Sheet ──
    from .scorer import score_config
    scoring = score_config(results)
    _add_executive_summary(wb, [{'filename': 'Config', 'config_type': config_type, 'scoring': scoring, 'versions': versions, 'results': results}])

    # ── Quick Reference Sheet ──
    ws = wb.create_sheet(title='Quick Reference')

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'].value = 'PAN-OS SD-WAN Feature Report'
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color=styles.BLUE)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:F2')
    ws['A2'].value = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}  |  Config Type: {config_type.upper()}'
    ws['A2'].font = styles.subtitle_font
    ws['A2'].alignment = Alignment(horizontal='center')

    # Group results by feature name (one row per source/device)
    result_groups = {}
    for r in results:
        if r.feature_name not in result_groups:
            result_groups[r.feature_name] = []
        result_groups[r.feature_name].append(r)

    # Quick reference by category — one row per device/source per feature
    row = 4
    headers = ['Category', 'Feature', 'Status', 'Device / Source', 'Configured Items', 'Count']
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    styles.style_header_row(ws, row, len(headers))
    row += 1

    for cat_name, features in FEATURE_CATEGORIES.items():
        cat_color = CAT_COLORS.get(cat_name, styles.BLUE)
        cat_fill = PatternFill(start_color=cat_color, end_color=cat_color, fill_type='solid')
        cat_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

        # Category header row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = cat_fill
            cell.font = cat_font
            cell.border = styles.thin_border
        ws.cell(row=row, column=1, value=cat_name)
        row += 1

        for feat_name in features:
            rlist = result_groups.get(feat_name)
            if rlist is None:
                # Feature parser not found — show as N/A
                ws.cell(row=row, column=1, value='')
                styles.style_data_cell(ws.cell(row=row, column=1), row)
                ws.cell(row=row, column=2, value=feat_name)
                styles.style_data_cell(ws.cell(row=row, column=2), row)
                cell = ws.cell(row=row, column=3)
                cell.value = 'N/A'
                cell.font = Font(name='Calibri', size=10, color='999999')
                cell.alignment = Alignment(horizontal='center')
                cell.border = styles.thin_border
                for col in range(4, len(headers) + 1):
                    styles.style_data_cell(ws.cell(row=row, column=col), row)
                row += 1
                continue

            # Separate enabled and disabled results
            enabled_results = [r for r in rlist if r.enabled]
            if not enabled_results:
                # All disabled — show one row
                ws.cell(row=row, column=1, value='')
                styles.style_data_cell(ws.cell(row=row, column=1), row)
                ws.cell(row=row, column=2, value=feat_name)
                styles.style_data_cell(ws.cell(row=row, column=2), row)
                styles.style_status_cell(ws.cell(row=row, column=3), False)
                ws.cell(row=row, column=4, value=rlist[0].source)
                styles.style_data_cell(ws.cell(row=row, column=4), row)
                ws.cell(row=row, column=5, value='Not configured')
                styles.style_data_cell(ws.cell(row=row, column=5), row)
                ws.cell(row=row, column=6, value='')
                styles.style_data_cell(ws.cell(row=row, column=6), row)
                row += 1
                continue

            # One row per enabled source/device
            for r in enabled_results:
                ws.cell(row=row, column=1, value='')
                styles.style_data_cell(ws.cell(row=row, column=1), row)
                ws.cell(row=row, column=2, value=feat_name)
                styles.style_data_cell(ws.cell(row=row, column=2), row)
                styles.style_status_cell(ws.cell(row=row, column=3), True)

                # Device/Source column
                ws.cell(row=row, column=4, value=r.source)
                styles.style_data_cell(ws.cell(row=row, column=4), row)

                # Configured Items — extract entry names from summary (format: "Source: Entry1, Entry2")
                items = r.summary
                if ':' in items:
                    items = items.split(':', 1)[1].strip()
                ws.cell(row=row, column=5, value=items)
                styles.style_data_cell(ws.cell(row=row, column=5), row)

                # Count
                count = len([e.strip() for e in items.split(',') if e.strip()]) if items else 0
                count_cell = ws.cell(row=row, column=6, value=count if count else '')
                count_cell.alignment = Alignment(horizontal='center')
                count_cell.border = styles.thin_border
                if count > 0:
                    count_cell.font = Font(name='Calibri', size=11, bold=True, color=styles.GREEN)
                else:
                    count_cell.font = styles.data_font
                row += 1

        # Blank row between categories
        row += 1

    # Enabled/Disabled totals
    row += 1
    enabled_count = sum(1 for r in results if r.enabled)
    disabled_count = sum(1 for r in results if not r.enabled)
    ws.cell(row=row, column=1, value='Total Features Configured:')
    ws.cell(row=row, column=1).font = Font(name='Calibri', size=11, bold=True)
    ws.cell(row=row, column=2, value=enabled_count)
    ws.cell(row=row, column=2).font = Font(name='Calibri', size=12, bold=True, color=styles.GREEN)
    ws.cell(row=row, column=3, value='Not Configured:')
    ws.cell(row=row, column=3).font = Font(name='Calibri', size=11, bold=True)
    ws.cell(row=row, column=4, value=disabled_count)
    ws.cell(row=row, column=4).font = Font(name='Calibri', size=12, bold=True, color=styles.RED)

    styles.auto_width(ws, min_width=12, max_width=55)
    ws.freeze_panes = 'A5'

    # ── Detail Sheets — merge same feature into one sheet ──
    # Group results by feature name
    from collections import OrderedDict
    grouped = OrderedDict()
    for result in results:
        if not result.rows or not result.columns:
            continue
        key = result.feature_name
        if key not in grouped:
            grouped[key] = {'columns': result.columns, 'rows': []}
        # Add source column to each row
        for data_row in result.rows:
            grouped[key]['rows'].append(list(data_row) + [result.source])

    seen_sheets = set(['Quick Reference'])
    for feat_name, data in grouped.items():
        sheet_name = feat_name
        for ch in '/\\?*[]:':
            sheet_name = sheet_name.replace(ch, '-')
        sheet_name = sheet_name[:31]
        if sheet_name in seen_sheets:
            sheet_name = sheet_name[:28] + f'_{len(seen_sheets)}'
        seen_sheets.add(sheet_name)

        ds = wb.create_sheet(title=sheet_name)

        # Headers — add Source column
        h_row = 1
        all_cols = data['columns'] + ['Source']
        for col, h in enumerate(all_cols, 1):
            ds.cell(row=h_row, column=col, value=h)
        styles.style_header_row(ds, h_row, len(all_cols))

        # Data rows
        for r_idx, data_row in enumerate(data['rows']):
            for c_idx, val in enumerate(data_row):
                cell = ds.cell(row=h_row + 1 + r_idx, column=c_idx + 1, value=val)
                styles.style_data_cell(cell, r_idx)

        # Auto-filter and width
        if data['rows']:
            last_col = get_column_letter(len(all_cols))
            ds.auto_filter.ref = f'A{h_row}:{last_col}{h_row + len(data["rows"])}'
        styles.auto_width(ds)
        ds.freeze_panes = f'A{h_row + 1}'

    # ── All Features Sheet (last) — split into Enabled and Disabled sections ──
    ws2 = wb.create_sheet(title='All Features')
    all_headers = ['Feature', 'Status', 'Summary', 'Source']

    enabled_results = [r for r in results if r.enabled]
    disabled_results = [r for r in results if not r.enabled]

    r2 = 1

    # Enabled section header
    enabled_color = '1E8449'
    enabled_section_fill = PatternFill(start_color=enabled_color, end_color=enabled_color, fill_type='solid')
    enabled_section_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    for col in range(1, len(all_headers) + 1):
        cell = ws2.cell(row=r2, column=col)
        cell.fill = enabled_section_fill
        cell.font = enabled_section_font
        cell.border = styles.thin_border
    ws2.cell(row=r2, column=1, value=f'Enabled Features ({len(enabled_results)})')
    r2 += 1

    # Enabled headers
    for col, h in enumerate(all_headers, 1):
        ws2.cell(row=r2, column=col, value=h)
    styles.style_header_row(ws2, r2, len(all_headers))
    r2 += 1

    # Enabled rows
    for result in enabled_results:
        ws2.cell(row=r2, column=1, value=result.feature_name)
        styles.style_data_cell(ws2.cell(row=r2, column=1), r2)
        styles.style_status_cell(ws2.cell(row=r2, column=2), result.enabled)
        ws2.cell(row=r2, column=3, value=result.summary)
        styles.style_data_cell(ws2.cell(row=r2, column=3), r2)
        ws2.cell(row=r2, column=4, value=result.source)
        styles.style_data_cell(ws2.cell(row=r2, column=4), r2)
        r2 += 1

    # Blank row separator
    r2 += 1

    # Disabled section header
    disabled_color = 'C0392B'
    disabled_section_fill = PatternFill(start_color=disabled_color, end_color=disabled_color, fill_type='solid')
    disabled_section_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    for col in range(1, len(all_headers) + 1):
        cell = ws2.cell(row=r2, column=col)
        cell.fill = disabled_section_fill
        cell.font = disabled_section_font
        cell.border = styles.thin_border
    ws2.cell(row=r2, column=1, value=f'Disabled / Not Configured ({len(disabled_results)})')
    r2 += 1

    # Disabled headers
    for col, h in enumerate(all_headers, 1):
        ws2.cell(row=r2, column=col, value=h)
    styles.style_header_row(ws2, r2, len(all_headers))
    r2 += 1

    # Disabled rows
    for result in disabled_results:
        ws2.cell(row=r2, column=1, value=result.feature_name)
        styles.style_data_cell(ws2.cell(row=r2, column=1), r2)
        styles.style_status_cell(ws2.cell(row=r2, column=2), result.enabled)
        ws2.cell(row=r2, column=3, value=result.summary)
        styles.style_data_cell(ws2.cell(row=r2, column=3), r2)
        ws2.cell(row=r2, column=4, value=result.source)
        styles.style_data_cell(ws2.cell(row=r2, column=4), r2)
        r2 += 1

    styles.auto_width(ws2)
    ws2.freeze_panes = 'A3'

    # Save
    filename = f'sdwan_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}_{uuid.uuid4().hex[:6]}.xlsx'
    out = output_dir or REPORT_DIR
    filepath = os.path.join(out, filename)
    wb.save(filepath)
    return filepath


def _sanitize_sheet_name(name, seen):
    """Sanitize a string for use as an Excel sheet name."""
    for ch in '/\\?*[]:':
        name = name.replace(ch, '-')
    name = name[:31]
    if name in seen:
        name = name[:28] + f'_{len(seen)}'
    seen.add(name)
    return name


def _aggregate_feature(results_list, feature_name):
    """Aggregate a feature across multiple results — enabled if any are enabled."""
    enabled = any(r.enabled for r in results_list)
    summaries = [r.summary for r in results_list if r.enabled and r.summary != 'Not configured']
    if not summaries:
        summaries = [r.summary for r in results_list]
    combined = '; '.join(dict.fromkeys(summaries))  # dedupe preserving order
    return enabled, combined


def generate_comparison(configs_data: list, output_dir: str = None) -> str:
    """Generate a comparison Excel report from multiple config files.

    Args:
        configs_data: list of dicts with keys: filename, config_type, results
    """
    from collections import OrderedDict
    from .scorer import score_configs

    wb = Workbook()

    # ── Executive Summary Sheet ──
    scored = score_configs(configs_data)
    _add_executive_summary(wb, scored)

    config_names = [c['filename'] for c in configs_data]
    num_configs = len(config_names)

    # Build per-config feature maps: {feature_name: aggregated (enabled, summary)}
    config_features = []
    for cfg in configs_data:
        feat_map = {}
        for r in cfg['results']:
            if r.feature_name not in feat_map:
                feat_map[r.feature_name] = []
            feat_map[r.feature_name].append(r)
        # Aggregate per feature
        aggregated = {}
        for fname, rlist in feat_map.items():
            enabled, summary = _aggregate_feature(rlist, fname)
            aggregated[fname] = {'enabled': enabled, 'summary': summary}
        config_features.append(aggregated)

    # ── Sheet 2: Comparison Summary ──
    ws = wb.create_sheet(title='Comparison Summary')

    # Title
    total_cols = 1 + num_configs * 2  # Feature + (Status, Summary) per config
    last_col_letter = get_column_letter(total_cols)
    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'].value = 'PAN-OS SD-WAN Configuration Comparison'
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color=styles.BLUE)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'A2:{last_col_letter}2')
    config_types = ', '.join(f"{c['filename']} ({c['config_type']})" for c in configs_data)
    ws['A2'].value = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}  |  Configs: {config_types}'
    ws['A2'].font = styles.subtitle_font
    ws['A2'].alignment = Alignment(horizontal='center')

    # Header row: Feature | Config1 Status | Config1 Summary | Config2 Status | ...
    row = 4
    headers = ['Feature']
    for name in config_names:
        short = name[:20] if len(name) > 20 else name
        headers.append(f'{short} Status')
        headers.append(f'{short} Summary')

    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    styles.style_header_row(ws, row, len(headers))
    row += 1

    # Feature rows grouped by category
    for cat_name, features in FEATURE_CATEGORIES.items():
        cat_color = CAT_COLORS.get(cat_name, styles.BLUE)
        cat_fill = PatternFill(start_color=cat_color, end_color=cat_color, fill_type='solid')
        cat_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = cat_fill
            cell.font = cat_font
            cell.border = styles.thin_border
        ws.cell(row=row, column=1, value=cat_name)
        row += 1

        for feat_name in features:
            ws.cell(row=row, column=1, value=feat_name)
            styles.style_data_cell(ws.cell(row=row, column=1), row)

            for cfg_idx in range(num_configs):
                feat_data = config_features[cfg_idx].get(feat_name)
                status_col = 2 + cfg_idx * 2
                summary_col = 3 + cfg_idx * 2

                if feat_data:
                    styles.style_status_cell(ws.cell(row=row, column=status_col), feat_data['enabled'])
                    ws.cell(row=row, column=summary_col, value=feat_data['summary'])
                    styles.style_data_cell(ws.cell(row=row, column=summary_col), row)
                else:
                    cell = ws.cell(row=row, column=status_col)
                    cell.value = 'N/A'
                    cell.font = Font(name='Calibri', size=10, color='999999')
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = styles.thin_border
                    styles.style_data_cell(ws.cell(row=row, column=summary_col), row)

            row += 1
        row += 1  # Blank row between categories

    # Totals per config
    row += 1
    ws.cell(row=row, column=1, value='Totals')
    ws.cell(row=row, column=1).font = Font(name='Calibri', size=11, bold=True)
    for cfg_idx in range(num_configs):
        status_col = 2 + cfg_idx * 2
        summary_col = 3 + cfg_idx * 2
        enabled = sum(1 for f in config_features[cfg_idx].values() if f['enabled'])
        disabled = sum(1 for f in config_features[cfg_idx].values() if not f['enabled'])
        ws.cell(row=row, column=status_col, value=f'{enabled} Enabled')
        ws.cell(row=row, column=status_col).font = Font(name='Calibri', size=11, bold=True, color=styles.GREEN)
        ws.cell(row=row, column=summary_col, value=f'{disabled} Disabled')
        ws.cell(row=row, column=summary_col).font = Font(name='Calibri', size=11, bold=True, color=styles.RED)

    styles.auto_width(ws, min_width=12, max_width=40)
    ws.freeze_panes = 'B5'

    # ── Detail Sheets — merge all configs per feature ──
    all_results_with_config = []
    for cfg in configs_data:
        for r in cfg['results']:
            all_results_with_config.append((cfg['filename'], r))

    grouped = OrderedDict()
    for cfg_name, result in all_results_with_config:
        if not result.rows or not result.columns:
            continue
        key = result.feature_name
        if key not in grouped:
            grouped[key] = {'columns': result.columns, 'rows': []}
        for data_row in result.rows:
            grouped[key]['rows'].append(list(data_row) + [result.source, cfg_name])

    seen_sheets = set(['Comparison Summary'])
    for feat_name, data in grouped.items():
        sheet_name = _sanitize_sheet_name(feat_name, seen_sheets)
        ds = wb.create_sheet(title=sheet_name)

        h_row = 1
        all_cols = data['columns'] + ['Source', 'Config File']
        for col, h in enumerate(all_cols, 1):
            ds.cell(row=h_row, column=col, value=h)
        styles.style_header_row(ds, h_row, len(all_cols))

        for r_idx, data_row in enumerate(data['rows']):
            for c_idx, val in enumerate(data_row):
                cell = ds.cell(row=h_row + 1 + r_idx, column=c_idx + 1, value=val)
                styles.style_data_cell(cell, r_idx)

        if data['rows']:
            last_col = get_column_letter(len(all_cols))
            ds.auto_filter.ref = f'A{h_row}:{last_col}{h_row + len(data["rows"])}'
        styles.auto_width(ds)
        ds.freeze_panes = f'A{h_row + 1}'

    # ── All Features Sheet (last) ──
    ws2 = wb.create_sheet(title='All Features')
    all_headers = ['Config File', 'Feature', 'Status', 'Summary', 'Source']

    all_flat = []
    for cfg in configs_data:
        for r in cfg['results']:
            all_flat.append((cfg['filename'], r))

    enabled_flat = [(c, r) for c, r in all_flat if r.enabled]
    disabled_flat = [(c, r) for c, r in all_flat if not r.enabled]

    r2 = 1

    # Enabled section
    enabled_section_fill = PatternFill(start_color='1E8449', end_color='1E8449', fill_type='solid')
    section_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    for col in range(1, len(all_headers) + 1):
        cell = ws2.cell(row=r2, column=col)
        cell.fill = enabled_section_fill
        cell.font = section_font
        cell.border = styles.thin_border
    ws2.cell(row=r2, column=1, value=f'Enabled Features ({len(enabled_flat)})')
    r2 += 1

    for col, h in enumerate(all_headers, 1):
        ws2.cell(row=r2, column=col, value=h)
    styles.style_header_row(ws2, r2, len(all_headers))
    r2 += 1

    for cfg_name, result in enabled_flat:
        ws2.cell(row=r2, column=1, value=cfg_name)
        styles.style_data_cell(ws2.cell(row=r2, column=1), r2)
        ws2.cell(row=r2, column=2, value=result.feature_name)
        styles.style_data_cell(ws2.cell(row=r2, column=2), r2)
        styles.style_status_cell(ws2.cell(row=r2, column=3), result.enabled)
        ws2.cell(row=r2, column=4, value=result.summary)
        styles.style_data_cell(ws2.cell(row=r2, column=4), r2)
        ws2.cell(row=r2, column=5, value=result.source)
        styles.style_data_cell(ws2.cell(row=r2, column=5), r2)
        r2 += 1

    r2 += 1

    # Disabled section
    disabled_section_fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
    for col in range(1, len(all_headers) + 1):
        cell = ws2.cell(row=r2, column=col)
        cell.fill = disabled_section_fill
        cell.font = section_font
        cell.border = styles.thin_border
    ws2.cell(row=r2, column=1, value=f'Disabled / Not Configured ({len(disabled_flat)})')
    r2 += 1

    for col, h in enumerate(all_headers, 1):
        ws2.cell(row=r2, column=col, value=h)
    styles.style_header_row(ws2, r2, len(all_headers))
    r2 += 1

    for cfg_name, result in disabled_flat:
        ws2.cell(row=r2, column=1, value=cfg_name)
        styles.style_data_cell(ws2.cell(row=r2, column=1), r2)
        ws2.cell(row=r2, column=2, value=result.feature_name)
        styles.style_data_cell(ws2.cell(row=r2, column=2), r2)
        styles.style_status_cell(ws2.cell(row=r2, column=3), result.enabled)
        ws2.cell(row=r2, column=4, value=result.summary)
        styles.style_data_cell(ws2.cell(row=r2, column=4), r2)
        ws2.cell(row=r2, column=5, value=result.source)
        styles.style_data_cell(ws2.cell(row=r2, column=5), r2)
        r2 += 1

    styles.auto_width(ws2)
    ws2.freeze_panes = 'A3'

    # Save
    filename = f'sdwan_comparison_{datetime.now().strftime("%Y%m%d_%H%M%S")}_{uuid.uuid4().hex[:6]}.xlsx'
    out = output_dir or REPORT_DIR
    filepath = os.path.join(out, filename)
    wb.save(filepath)
    return filepath
